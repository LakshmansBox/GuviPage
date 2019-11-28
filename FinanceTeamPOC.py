import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

login = os.getlogin()
pipeini = 'C:/Users/'+login+'/Downloads/Pipeline'
piperev = 'C:/Users/'+login+'/Downloads/Pipeline_revised'

alpha_dic = {}

def char_range(c1, c2):
    """Generates the characters from `c1` to `c2`, inclusive."""
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)

def get_colname():
    count=1
    for a in char_range('A', 'Z'):
        alpha_dic[count] = a
        count+=1

    count=27
    for a in char_range('A', 'Z'):
        alpha_dic[count] = "A"+a
        count+=1

    count=53
    for a in char_range('A', 'Z'):
        alpha_dic[count] = "B"+a
        count+=1

    return alpha_dic

alpha_dic = get_colname()




def modify(piperev,file, finallist, valuelist, final_del,len_of_rb2):
    wb = openpyxl.load_workbook(os.path.join(piperev, file))
    sheet = wb['Pipeline']
    redfill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    redfill2 = PatternFill(fill_type='solid', start_color='FF3333', end_color='FF4444')
    count = 0
    for i in finallist:
        count+=1
        if i!="Match_Found":
            gen_list = [(str(x)+str(count+1)) for x in valuelist]
            for cell in gen_list:
                sheet[cell].fill = redfill
    del_count = 0   
    j = 0         
    for item in final_del:
        j +=1
        if item == "no":
            del_count+=1
            no = len_of_rb2+ del_count+1
            sheet["A"+ str(no)] = "Row "+str(j)+ " is removed/modified"
            sheet["A"+ str(no)].fill = redfill2 
            
    wb.save(os.path.join(piperev, file))
    
    
def check_deleted(rb1, rb2):
    final_del = []
    for row in range(len(rb1)):
            deleted_info =[]
            del_df1 = rb1[rb1['index']==row].copy()
            del_df1 = del_df1.drop(labels=['index'], axis=1)
            del_df1 = del_df1.reset_index(drop=True)
            for row in range(len(rb2)):
                del_df2 = rb2[rb2['index']==row].copy()
                del_df2 = del_df2.drop(labels=['index'], axis=1)
                del_df2 = del_df2.reset_index(drop=True)
                if del_df1.equals(del_df2):
                    deleted_info.append("yes")
                else:
                    deleted_info.append("no")
                    
            if "yes" in deleted_info:
                final_del.append("yes")
            else:
                final_del.append("no")
            
    return final_del
    
        
        
def count_files(pipeini, piperev, alphadic):
    print("Analysing Files.........................\n")
    print("Donot Press any key till Processing\n")
    filelist = os.listdir(piperev)
    
    for file in filelist:
        
        rb1 = pd.read_excel(os.path.join(pipeini, file))
        rb2 = pd.read_excel(os.path.join(piperev, file))
        rb1 = rb1.reset_index()
        rb2 = rb2.reset_index()
        
        list2 = []
        for row in range(len(rb2)):
            list1 =[]
            temp_df = rb2[rb2['index']==row].copy()
            temp_df = temp_df.drop(labels=['index'], axis=1)
            temp_df = temp_df.reset_index(drop=True)
            for row in range(len(rb1)):
                temp_df2 = rb1[rb1['index']==row].copy()
                temp_df2 = temp_df2.drop(labels=['index'], axis=1)
                temp_df2 = temp_df2.reset_index(drop=True)
                if temp_df.equals(temp_df2):
                    list1.append("yes")
                else:
                    list1.append("no")
            
            if "yes" in list1:
                list2.append("Match_Found")
            else:
                list2.append("Match_Not_Found")
            finallist = list2
            valuelist = list(alpha_dic.values())
            len_of_rb2 = len(rb2)
        final_del  = check_deleted(rb1, rb2)
        
        print("Highlighting the Differences in "+file)
        modify(piperev, file, finallist, valuelist, final_del, len_of_rb2)
        print("Completed "+file+"\n")

count_files(pipeini, piperev, alpha_dic)
print("Process Completed Successfully\n")
input("Press Enter to Exit...............")
        
    
        
        

        
    
